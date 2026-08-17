"""Read-only access to a local .xlsx export of the order sheet — same pivot
as sheets_source.py (wide per-person qty matrix -> one label-record per
(person, part) pair where qty > 0), but locating columns by header name
instead of fixed indices.

Why not reuse sheets_source's fixed COL_* indices? A downloaded/exported
.xlsx of the sheet is not guaranteed to have the same column layout as the
live Google Sheet (extra leading columns, renamed headers, etc. have been
observed in practice) — matching by header text is robust to that drift.

READ-ONLY: only ever reads the workbook, never writes to it.
"""

import sys
from openpyxl import load_workbook

import config
from sheets_source import LabelRecord, SheetIssue

# Header text candidates, in priority order, for each fixed column. The
# first match found in the header row wins.
ELEMENT_ID_HEADERS = ("Element ID", "Part Number")
DESCRIPTION_HEADERS = ("Description",)
COLOR_HEADERS = ("BL Color", "LEGO Color", "Color")


def _find_col(header: list, candidates: tuple[str, ...]) -> int | None:
    lower = [str(h).strip().lower() if h is not None else "" for h in header]
    for name in candidates:
        target = name.lower()
        if target in lower:
            return lower.index(target)
    return None


def _find_col_with_data(
    header: list, candidates: tuple[str, ...], data_rows: list[list]
) -> int | None:
    """Like _find_col, but among matching candidate headers prefers one
    whose column actually has data — an export has been seen to carry a
    blank "BL Color" column alongside a populated "LEGO Color" one."""
    matches = []
    lower = [str(h).strip().lower() if h is not None else "" for h in header]
    for name in candidates:
        target = name.lower()
        if target in lower:
            matches.append(lower.index(target))
    if not matches:
        return None
    for col in matches:
        for row in data_rows[:50]:
            if col < len(row) and row[col] not in (None, ""):
                return col
    return matches[0]  # all candidates blank in the sample — fall back to first


def _resolve_tab_name(tab: str, sheetnames: list[str]) -> str | None:
    """Exact match first; otherwise fall back to a whitespace/case-insensitive
    match — an .xlsx export of the sheet has been seen to drop spaces from
    tab names (e.g. "Order Here" -> "OrderHere")."""
    if tab in sheetnames:
        return tab
    normalized = tab.replace(" ", "").lower()
    for name in sheetnames:
        if name.replace(" ", "").lower() == normalized:
            return name
    return None


def _load_rows(path: str, tab: str) -> list[list]:
    try:
        wb = load_workbook(path, data_only=True, read_only=True)
    except FileNotFoundError:
        sys.exit(f"Source file not found: '{path}'")
    except (KeyError, OSError) as e:
        sys.exit(f"Couldn't open '{path}' as an Excel workbook: {e}")

    resolved = _resolve_tab_name(tab, wb.sheetnames)
    if resolved is None:
        sys.exit(
            f"Tab '{tab}' not found in '{path}'. Sheets in this file: "
            f"{', '.join(wb.sheetnames)}"
        )
    ws = wb[resolved]
    return [[c for c in row] for row in ws.iter_rows(values_only=True)]


def _cell(row: list, idx: int | None, default=""):
    if idx is None or idx >= len(row):
        return default
    val = row[idx]
    if val is None or val == "":
        return default
    return val


def _find_header_row(rows: list[list]) -> int | None:
    """config.HEADER_ROW is tuned for the live Google Sheet's layout; an
    .xlsx export has been seen to have extra rows above the real header
    (e.g. a totals row). Try the configured row first, then scan nearby
    rows for one that actually has a recognizable element-id column."""
    search_order = [config.HEADER_ROW, *range(0, min(len(rows), 10))]
    seen = set()
    for row_idx in search_order:
        if row_idx in seen or row_idx >= len(rows):
            continue
        seen.add(row_idx)
        if _find_col(rows[row_idx], ELEMENT_ID_HEADERS) is not None:
            return row_idx
    return None


def _build_records(rows: list[list]) -> tuple[list[LabelRecord], list[SheetIssue]]:
    if not rows:
        return [], []

    header_row = _find_header_row(rows)
    if header_row is None:
        sys.exit(
            f"Couldn't find a header row with an element ID column "
            f"(looked for {ELEMENT_ID_HEADERS}) in the first 10 rows."
        )
    header = rows[header_row]
    # Data starts however many rows below the header the live sheet's
    # DATA_START_ROW/HEADER_ROW gap implies (usually 1 blank/units row).
    data_start_row = header_row + (config.DATA_START_ROW - config.HEADER_ROW)

    data_rows = rows[data_start_row:]
    col_element_id = _find_col(header, ELEMENT_ID_HEADERS)
    col_description = _find_col_with_data(header, DESCRIPTION_HEADERS, data_rows)
    col_color = _find_col_with_data(header, COLOR_HEADERS, data_rows)

    # Person columns: a contiguous run of (name, cost) column pairs — a text
    # header immediately followed by a *numeric* header (the per-person
    # running total/cost cell, e.g. 232.45) — starting right after the
    # fixed columns. Other single text-header metadata columns (e.g. "BL
    # Price", "Nominated for", the odd standalone "Price" column) are
    # skipped since they aren't followed by a numeric header. Stop at the
    # first pair that doesn't match rather than scanning to the end of the
    # sheet — real exports have been seen to have unrelated debris (broken
    # formula refs, other tables) further to the right.
    scan_start = max(
        c for c in (col_element_id, col_description, col_color) if c is not None
    ) + 1

    def _is_name_cost_pair(col: int) -> bool:
        if col + 1 >= len(header):
            return False
        name, cost = header[col], header[col + 1]
        return (
            isinstance(name, str) and name.strip()
            and isinstance(cost, (int, float))
        )

    first_person_col = scan_start
    while first_person_col < len(header) and not _is_name_cost_pair(first_person_col):
        first_person_col += 1

    person_cols = []
    col = first_person_col
    while col < len(header) and _is_name_cost_pair(col):
        person_cols.append((col, header[col].strip()))
        col += 2

    records: list[LabelRecord] = []
    issues: list[SheetIssue] = []
    seen: set[tuple[str, str]] = set()

    for offset, row in enumerate(rows[data_start_row:]):
        sheet_row = data_start_row + offset + 1
        if not row:
            continue
        raw_id = _cell(row, col_element_id)
        if raw_id == "":
            continue  # blank/footer row
        element_id = _format_element_id(raw_id)

        description = str(_cell(row, col_description))
        if not description:
            issues.append(SheetIssue(sheet_row, "missing_description",
                                      f"Element {element_id} has no description"))

        color = config.COLOR_OVERRIDES.get(element_id, str(_cell(row, col_color)))
        if not color:
            issues.append(SheetIssue(sheet_row, "missing_color",
                                      f"Element {element_id} has no color"))

        image_url = config.IMAGE_URL_TEMPLATE.format(element_id=element_id)

        for qty_col, person in person_cols:
            raw_qty = _cell(row, qty_col)
            if raw_qty == "":
                continue  # blank cell, not a mistake
            qty = str(raw_qty).strip()
            if not qty:
                continue
            try:
                qty_num = float(qty.replace(",", ""))
            except ValueError:
                issues.append(SheetIssue(
                    sheet_row, "bad_qty",
                    f"{person}'s qty for element {element_id} is non-numeric: '{qty}'",
                ))
                continue
            if qty_num <= 0:
                continue

            key = (person, element_id)
            if key in seen:
                issues.append(SheetIssue(
                    sheet_row, "duplicate",
                    f"{person} has more than one qty entry for element {element_id}",
                ))
            seen.add(key)

            # Keep qty formatted the way sheets_source does (string, may
            # include a trailing ".0" from Excel's numeric cells is trimmed).
            records.append(
                LabelRecord(
                    person=person,
                    element_id=element_id,
                    description=description,
                    color=color,
                    qty=_format_qty(qty_num),
                    image_url=image_url,
                )
            )

    return records, issues


def _format_element_id(raw) -> str:
    if isinstance(raw, float) and raw.is_integer():
        return str(int(raw))
    return str(raw).strip()


def _format_qty(qty_num: float) -> str:
    if qty_num.is_integer():
        return str(int(qty_num))
    return str(qty_num)


def get_label_records(
    path: str, tab: str = config.SOURCE_TAB
) -> list[LabelRecord]:
    rows = _load_rows(path, tab)
    records, _issues = _build_records(rows)
    return records


def validate_source(
    path: str, tab: str = config.SOURCE_TAB
) -> tuple[list[LabelRecord], list[SheetIssue]]:
    rows = _load_rows(path, tab)
    return _build_records(rows)
